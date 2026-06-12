import streamlit as st
from tools import get_patient_history, update_medical_record, book_appointment, search_disease_info
from agent import agent_executor

st.set_page_config(page_title="Agentic Healthcare Assistant", page_icon="🏥", layout="wide")
st.title("🏥 Agentic Healthcare Assistant")
st.markdown("Your AI-powered virtual medical assistant")

st.sidebar.title("Navigation")
page = st.sidebar.radio("Select a page", ["Chat Assistant", "Patient Records", "Book Appointment", "Medical Search"])

if page == "Chat Assistant":
    st.header("Chat with your Healthcare Assistant")
    if "messages" not in st.session_state:
        st.session_state.messages = []
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
    if prompt := st.chat_input("Ask me anything about your health..."):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)
        with st.chat_message("assistant"):
            with st.spinner("Thinking..."):
                prompt_lower = prompt.lower()
                # Get patient history
                if "history" in prompt_lower or "record" in prompt_lower:
                    from openpyxl import load_workbook
                    wb = load_workbook("Agentic Healthcare Assistant for Medical Task Automation/records.xlsx")
                    ws = wb['Sheet1']
                    names = [row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[2]]
                    found = False
                    for name in names:
                        if name.lower() in prompt_lower:
                            result = get_patient_history(name)
                            answer = f"**Patient:** {result['name']}\n\n**Age:** {result['age']}\n\n**Gender:** {result['gender']}\n\n**Address:** {result['address']}\n\n**Summary:** {result['summary'] or 'No summary available'}"
                            found = True
                            break
                    if not found:
                        answer = "Please mention the patient's full name in your query."
                # Book appointment
                elif "book" in prompt_lower or "appointment" in prompt_lower:
                    answer = "Please use the **Book Appointment** page from the sidebar to schedule an appointment."
                # Medical search
                elif "disease" in prompt_lower or "treatment" in prompt_lower or "search" in prompt_lower:
                    result = search_disease_info(prompt)
                    answer = result["results"]
                else:
                    answer = "Hello! I'm your Healthcare Assistant. I can help you with:\n\n- **Patient history** — e.g. 'Show Rahul Negi history'\n- **Book appointment** — use the Book Appointment page\n- **Medical search** — e.g. 'Search kidney disease treatment'"
                st.markdown(answer)
        st.session_state.messages.append({"role": "assistant", "content": answer})

elif page == "Patient Records":
    st.header("Patient Records")
    name = st.text_input("Enter Patient Name")
    if st.button("Get Patient History"):
        if name:
            result = get_patient_history(name)
            if "error" in result:
                st.error(result["error"])
            else:
                st.success("Patient Found!")
                col1, col2 = st.columns(2)
                col1.metric("Name", result["name"])
                col1.metric("Age", result["age"])
                col2.metric("Gender", result["gender"])
                col2.metric("Address", result["address"])
                st.subheader("Medical Summary")
                st.info(result["summary"] if result["summary"] else "No summary available")
        else:
            st.warning("Please enter a patient name")

elif page == "Book Appointment":
    st.header("Book an Appointment")
    patient_name = st.text_input("Patient Name")
    doctor_speciality = st.selectbox("Doctor Speciality", [
        "General Physician", "Nephrologist", "Cardiologist",
        "Neurologist", "Orthopedist", "Dermatologist"
    ])
    date = st.date_input("Appointment Date")
    if st.button("Book Appointment"):
        if patient_name:
            result = book_appointment(patient_name, doctor_speciality, str(date))
            if "error" in result:
                st.error(result["error"])
            else:
                st.success(result["message"])
                st.json(result)
        else:
            st.warning("Please enter a patient name")

elif page == "Medical Search":
    st.header("Medical Information Search")
    query = st.text_input("Search for a disease or treatment")
    if st.button("Search"):
        if query:
            with st.spinner("Searching..."):
                result = search_disease_info(query)
                st.subheader("Results")
                st.info(result["results"])
        else:
            st.warning("Please enter a search query")