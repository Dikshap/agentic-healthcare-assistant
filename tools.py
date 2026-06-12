import openpyxl
from rag_pipeline import retrieve

EXCEL_PATH = "Agentic Healthcare Assistant for Medical Task Automation/records.xlsx"

#get_patient_history
def get_patient_history(name):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] and row[2].lower() == name.lower():
            return {
                "name": row[2],
                "age": row[3],
                "gender": row[4],
                "address": row[5],
                "summary": row[6]
            }
    return {"error": f"No patient found with name {name}"}

#update_medical_record
def update_medical_record(name, summary):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=2):
        if row[2].value and row[2].value.lower() == name.lower():
            row[6].value = summary
            wb.save(EXCEL_PATH)
            return {"status": f"Record updated successfully for {name}"}
    return {"error": f"No patient found with name {name}"}

#book_appointment
def book_appointment(patient_name, doctor_speciality, date):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] and row[2].lower() == patient_name.lower():
            return {
                "status": "Appointment booked successfully",
                "patient": patient_name,
                "doctor_speciality": doctor_speciality,
                "date": date,
                "message": f"Appointment booked for {patient_name} with a {doctor_speciality} on {date}"
            }
    return {"error": f"No patient found with name {patient_name}"}

#search_diesease_info
def search_disease_info(query):
    results = retrieve(query)
    return {"query": query, "results": results}

if __name__ == "__main__":
    print(get_patient_history("Rahul Negi"))
    print(book_appointment("Rahul Negi", "nephrologist", "2024-12-01"))
    print(update_medical_record("Rahul Negi", "Patient has chronic kidney disease"))
    print(search_disease_info("kidney disease treatment"))